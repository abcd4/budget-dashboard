import pandas as pd
import sqlite3
from openpyxl import load_workbook

# ── Load the workbook ──────────────────────────────────────
# read_only=True loads it faster and uses less memory
def load_all_data():
  wb = load_workbook("findata.xlsx", read_only=True)

  # ── Empty lists to collect rows across all sheets ──────────
  # We'll append to these as we loop through each month
  all_income = []
  all_expenses = []
  all_budget = []

  # ── Sheets to process ─────────────────────────────────────
  sheets = ["FEB 26", "MAR 26", "APR 26"]

  for sheet in sheets:
      ws = wb[sheet]
      
      # Convert the sheet into a list of tuples (one per row)
      rows = list(ws.iter_rows(values_only=True))

      # Skip row 0 (headers) and loop through the rest
      for row in rows[1:]:

          # ── INCOME (columns A–D = indexes 0–3) ────────────
          # Only grab rows where amount is a number and description is text
          # Skip the "TOTAL CASH FLOW" summary row
          desc, amount, cat, date = row[0], row[1], row[2], row[3]
          if isinstance(amount, (int, float)) and isinstance(desc, str) and desc not in ("TOTAL CASH FLOW",):
              all_income.append({
                  "description": desc,
                  "amount": amount,
                  "category": cat,
                  "date": date,
                  "month": sheet  # tag each row with its month so we can filter later
              })

          # ── EXPENSES (columns I–L = indexes 8–11) ─────────
          # Same logic — only grab rows with a real number and a description
          desc, amount, cat, date = row[8], row[9], row[10], row[11]
          if isinstance(amount, (int, float)) and isinstance(desc, str):
              all_expenses.append({
                  "description": desc,
                  "amount": amount,
                  "category": cat,
                  "date": date,
                  "month": sheet
              })

          # ── BUDGET TARGETS (columns P, S = indexes 15, 18) ─
          # Check row length first — some rows are shorter and
          # accessing index 18 on them would crash the script
          if len(row) > 18:
              cat, ideal = row[15], row[18]
              # Skip the summary total row, keep everything else
              # even if ideal_budget is None (no budget set yet)
              if isinstance(cat, str) and cat != "Expense Category Total":
                  all_budget.append({
                      "category": cat,
                      "ideal_budget": ideal,  # can be None — that's ok
                      "month": sheet
                  })

  # ── Load everything into SQLite ────────────────────────────
  # if_exists="replace" wipes and rewrites the table each time
  # so running this script again always gives you fresh data
  conn = sqlite3.connect("budget.db")
  pd.DataFrame(all_income).to_sql("income", conn, if_exists="replace", index=False)
  pd.DataFrame(all_expenses).to_sql("expenses", conn, if_exists="replace", index=False)
  pd.DataFrame(all_budget).to_sql("budget_targets", conn, if_exists="replace", index=False)

  print("✅ Income rows:", len(all_income))
  print("✅ Expense rows:", len(all_expenses))
  print("✅ Budget rows:", len(all_budget))

  conn.close()